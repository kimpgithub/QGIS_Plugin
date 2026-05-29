"""명부(행정리현황) 엑셀 파서.

대전 작업자의 로컬 작업목록. 발주처가 엑셀로 제공 — 어떤 행정리를 그려야
하는지의 마스터 리스트. 행정리 작업 탭에서 read_excel 로 읽어 표시한다.

정본 엑셀 스키마 — (2025농총) 행정리현황 명부, 11컬럼:
    CTPV_CD, CTPV_NM, SGG_CD, SGG_NM, EMD_CD, EMD_NM,
    RI_NM, LI_NM, LI_CD, WORK_YN, REMARK

주의 — 이 명부는 RI_/LI_ 의미가 플러그인 내부 표준과 뒤바뀌어 있다:
    엑셀 RI_NM = 법정리명  → 내부 li_nm
    엑셀 LI_NM = 행정리명  → 내부 ri_nm
    엑셀 LI_CD = 행정리코드 → 내부 ri_cd
내부 표준(adm=읍면동, ri=행정리)에 맞춰 매핑한다.
"""
import os
import re


# 엑셀 헤더 → 내부 표준 컬럼명 매핑.
# 정본은 (2025농총) 명부 — CTPV_CD/EMD_CD/LI_CD 계열. 헤더는 _normalize 로
# 비교(대소문자/공백/언더바 무시)한다.
COLUMN_ALIASES = {
    'sido_cd':    ['ctpv_cd', 'sido_cd', '시도코드'],
    'sido_nm':    ['ctpv_nm', 'sido_nm', '시도명', '시도명칭'],
    'sigungu_cd': ['sgg_cd', 'sigungu_cd', '시군구코드'],
    'sigungu_nm': ['sgg_nm', 'sigungu_nm', '시군구명', '시군구명칭'],
    'adm_cd':     ['emd_cd', 'adm_cd', '읍면동코드', '행정동코드'],
    'adm_nm':     ['emd_nm', 'adm_nm', '읍면동명', '행정동명', '행정동명칭'],
    'li_nm':      ['ri_nm', '법정리명', '리명'],          # 명부 RI_NM = 법정리명
    'ri_nm':      ['li_nm', '행정리명', '행정리명칭'],     # 명부 LI_NM = 행정리명
    'ri_cd':      ['li_cd', '행정리코드'],                # 명부 LI_CD = 행정리코드
    'work_yn':    ['work_yn', '작업여부', '작업완료여부'],
    'remark':     ['remark', 'rmk', '비고', '특이사항'],
}

REQUIRED_COLS = ['adm_cd', 'adm_nm', 'ri_cd', 'ri_nm']

# read_excel 결과 row dict 에 항상 존재하는 키 (없으면 '')
ALL_COLS = ['sido_cd', 'sido_nm', 'sigungu_cd', 'sigungu_nm',
            'adm_cd', 'adm_nm', 'li_nm', 'ri_nm', 'ri_cd',
            'work_yn', 'remark']


def _normalize(s):
    """컬럼명 정규화 — 공백 제거 + 소문자 + 특수문자 제거."""
    if s is None:
        return ''
    return re.sub(r'[\s_\-]+', '', str(s).strip().lower())


def detect_columns(header_row):
    """엑셀 헤더 → 표준 컬럼명 매핑. 매핑 안 된 컬럼은 무시.

    Returns:
        {엑셀_헤더: 표준_컬럼} dict, 누락된 required 컬럼 list
    """
    mapping = {}
    for h in header_row:
        hn = _normalize(h)
        if not hn:
            continue
        for canonical, aliases in COLUMN_ALIASES.items():
            if hn in {_normalize(a) for a in aliases}:
                mapping[str(h)] = canonical
                break
    found = set(mapping.values())
    missing = [c for c in REQUIRED_COLS if c not in found]
    return mapping, missing


def _pick_sheet(wb):
    """헤더에 required 컬럼이 다 있는 시트를 고른다.

    명부 엑셀은 '행정리현황' + '테이블정의' 두 시트가 있고 active 가 후자를
    가리킬 수 있어 wb.active 를 믿지 않는다. 매칭 시트가 없으면 첫 시트.
    """
    for ws in wb.worksheets:
        first = next(ws.iter_rows(values_only=True), None)
        if not first:
            continue
        hdr = [str(h) if h is not None else '' for h in first]
        _mapping, missing = detect_columns(hdr)
        if not missing:
            return ws
    return wb.worksheets[0] if wb.worksheets else None


def update_work_yn(path, adm_cd, ri_cd, value, sheet=None):
    """명부 엑셀에서 (adm_cd, ri_cd) 일치 행의 WORK_YN 값을 갱신.

    명부 ri_cd 는 읍면동 안에서만 유일(001/002…)하므로 반드시 adm_cd 와
    함께 매칭해야 한다. 첫 호출 시 {path}.bak 백업 자동 생성.
    매칭 행 수(int) 반환. 다른 프로세스가 파일을 열고 있으면 PermissionError.
    """
    import openpyxl
    import shutil
    bak = path + '.bak'
    if not os.path.exists(bak):
        try:
            shutil.copy2(path, bak)
        except Exception:
            pass
    wb = openpyxl.load_workbook(path)
    try:
        ws = wb[sheet] if sheet else _pick_sheet(wb)
        if ws is None:
            return 0
        rows_iter = ws.iter_rows()
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return 0
        headers = [str(c.value) if c.value is not None else ''
                   for c in header_row]
        mapping, _missing = detect_columns(headers)
        canon_to_idx = {}
        for i, h in enumerate(headers):
            canon = mapping.get(h)
            if canon:
                canon_to_idx[canon] = i
        idx_adm = canon_to_idx.get('adm_cd')
        idx_ri = canon_to_idx.get('ri_cd')
        idx_yn = canon_to_idx.get('work_yn')
        if idx_adm is None or idx_ri is None or idx_yn is None:
            return 0
        t_adm = str(adm_cd).strip()
        t_ri = str(ri_cd).strip()
        new_val = str(value or '').strip()
        n = 0
        for row in rows_iter:
            cv_adm = row[idx_adm].value
            cv_ri = row[idx_ri].value
            if cv_adm is None or cv_ri is None:
                continue
            if (str(cv_adm).strip() == t_adm
                    and str(cv_ri).strip() == t_ri):
                row[idx_yn].value = new_val
                n += 1
        if n:
            wb.save(path)
        return n
    finally:
        wb.close()


class WorkbookEditor:
    """명부 워크북을 1회만 로딩해 메모리에 유지하며 WORK_YN 셀을 수정/일괄 저장.

    38K행 명부는 load≈3s·save≈5s 라 변경마다 load+save 하던 update_work_yn 은
    건당 ~8s 가 걸린다. 이 클래스는 load 1회 + (adm_cd, ri_cd)→행번호 색인 1회로
    셀 수정을 즉시(수 ms) 처리하고, 저장은 flush() 호출 시 1회만 한다.

    openpyxl 은 스레드세이프하지 않다 — 반드시 단일 스레드(저장 워커)에서만 쓸 것.
    """

    def __init__(self, path, sheet=None):
        self.path = path
        self.sheet = sheet
        self._wb = None
        self._ws = None
        self._col_idx = {}       # canonical 컬럼명 → 0-base 열 인덱스
        self._row_of = {}        # (adm_cd, ri_cd) → [엑셀 행번호…]
        self._dirty = False

    def _ensure_loaded(self):
        if self._wb is not None:
            return
        import openpyxl
        import shutil
        bak = self.path + '.bak'
        if not os.path.exists(bak):
            try:
                shutil.copy2(self.path, bak)
            except Exception:
                pass
        self._wb = openpyxl.load_workbook(self.path)
        self._ws = (self._wb[self.sheet] if self.sheet
                    else _pick_sheet(self._wb)) or self._wb.active
        rows_iter = self._ws.iter_rows()
        header_row = next(rows_iter, None)
        if header_row is None:
            return
        headers = [str(c.value) if c.value is not None else ''
                   for c in header_row]
        mapping, _missing = detect_columns(headers)
        for i, h in enumerate(headers):
            canon = mapping.get(h)
            if canon:
                self._col_idx[canon] = i
        idx_adm = self._col_idx.get('adm_cd')
        idx_ri = self._col_idx.get('ri_cd')
        if idx_adm is None or idx_ri is None:
            return
        for cells in rows_iter:
            cv_adm = cells[idx_adm].value
            cv_ri = cells[idx_ri].value
            if cv_adm is None or cv_ri is None:
                continue
            key = (str(cv_adm).strip(), str(cv_ri).strip())
            self._row_of.setdefault(key, []).append(cells[0].row)

    def set_cell(self, adm_cd, ri_cd, field, value):
        """(adm_cd, ri_cd) 매칭 행의 canonical 컬럼 1개를 메모리에서 즉시 갱신.

        field: 'work_yn' | 'remark' 등 COLUMN_ALIASES canonical 키.
        매칭 행 수 반환. 디스크 저장은 flush(). 해당 컬럼이 없으면 0.
        """
        self._ensure_loaded()
        col = self._col_idx.get(field)
        if col is None:
            return 0
        rows = self._row_of.get((str(adm_cd).strip(), str(ri_cd).strip()), [])
        new_val = str(value if value is not None else '').strip()
        for rownum in rows:
            self._ws.cell(row=rownum, column=col + 1).value = new_val
        if rows:
            self._dirty = True
        return len(rows)

    def set_work_yn(self, adm_cd, ri_cd, value):
        """하위호환 — set_cell(..., 'work_yn', ...)."""
        return self.set_cell(adm_cd, ri_cd, 'work_yn', value)

    def flush(self):
        """변경분이 있으면 디스크에 1회 저장. 저장했으면 True.

        파일이 다른 프로세스에 열려 있으면 PermissionError 가 전파된다.
        """
        if self._wb is None or not self._dirty:
            return False
        self._wb.save(self.path)
        self._dirty = False
        return True

    def close(self):
        if self._wb is not None:
            try:
                self._wb.close()
            finally:
                self._wb = None


def read_excel(path, sheet=None, limit=None, adm_codes=None):
    """엑셀 파일 → (headers, rows, mapping, missing).

    - headers: 원본 엑셀 헤더 list
    - rows: 데이터 행 (원본 순서 유지) — dict with canonical keys
    - mapping: 엑셀 헤더 → 표준 컬럼
    - missing: 누락된 required 컬럼
    - adm_codes: 주어지면 해당 set 의 adm_cd 행만 읽어들임 (38K 행 명부에서
      병합이미지 대상 읍면동만 추출하는 용도)
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else _pick_sheet(wb)
    if ws is None:
        wb.close()
        return [], [], {}, REQUIRED_COLS.copy()
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return [], [], {}, REQUIRED_COLS.copy()
    headers = [str(h) if h is not None else '' for h in header_row]
    mapping, missing = detect_columns(headers)

    # 엑셀 헤더 → canonical 이름으로 변환
    col_to_canon = {i: mapping[headers[i]] for i in range(len(headers))
                    if headers[i] in mapping}
    # adm_cd 컬럼 인덱스 (사전 필터에 사용)
    adm_idx = next((i for i, c in col_to_canon.items()
                    if c == 'adm_cd'), None)
    code_set = (set(str(c).strip() for c in adm_codes)
                if adm_codes else None)
    rows = []
    for r in rows_iter:
        if all(c is None for c in r):
            continue
        # adm_cd 사전 필터 — dict 빌드 전에 조기 컷
        if code_set is not None and adm_idx is not None:
            v = r[adm_idx] if adm_idx < len(r) else None
            if v is None or str(v).strip() not in code_set:
                continue
        rec = {}
        for i, v in enumerate(r):
            if i in col_to_canon:
                if v is None:
                    rec[col_to_canon[i]] = ''
                else:
                    rec[col_to_canon[i]] = str(v).strip()
        for c in ALL_COLS:
            rec.setdefault(c, '')
        if not rec['adm_cd'] or not rec['ri_cd'] or not rec['ri_nm']:
            continue
        rows.append(rec)
        if limit and len(rows) >= limit:
            break
    wb.close()
    return headers, rows, mapping, missing
